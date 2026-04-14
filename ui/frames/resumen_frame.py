# ui/frames/resumen_frame.py

import tkinter as tk
from tkinter import ttk, messagebox
from database.queries import (
    obtener_meses,
    obtener_mes_actual,
    obtener_resumen_mes,
    obtener_desglose_por_categoria,
)
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from database.queries import obtener_movimientos

MESES_NOMBRES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

COLOR_INGRESO = "#27ae60"
COLOR_EGRESO  = "#e74c3c"
COLOR_NEUTRO  = "#2c3e50"


class ResumenFrame(tk.Frame):

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.configure(bg="#f0f2f5")

        self.mes_actual = None
        self._meses     = []

        self._build_ui()
        self._cargar_meses()

    # ─── Construcción de la UI ────────────────────────────────

    def _build_ui(self):
        # ── Encabezado con selector de mes ──
        header = tk.Frame(self, bg="#2c3e50", pady=12)
        header.pack(fill="x")

        tk.Label(header, text="Resumen mensual", font=("Arial", 16, "bold"),
                 bg="#2c3e50", fg="white").pack(side="left", padx=20)

        self.combo_mes = ttk.Combobox(header, state="readonly", width=20)
        self.combo_mes.pack(side="right", padx=20)
        self.combo_mes.bind("<<ComboboxSelected>>", self._on_mes_seleccionado)

        tk.Button(
            header, text="⬇ Exportar Excel",
            bg="#27ae60", fg="white",
            font=("Arial", 10, "bold"),
            bd=0, padx=10, pady=4,
            cursor="hand2",
            command=self._exportar_excel
        ).pack(side="right", padx=(0, 10))

        # ── Tarjetas de resumen ──
        self.frame_tarjetas = tk.Frame(self, bg="#f0f2f5")
        self.frame_tarjetas.pack(fill="x", padx=20, pady=(20, 10))

        self.lbl_saldo_inicial = self._tarjeta(self.frame_tarjetas, "Saldo inicial",  "#8e44ad")
        self.lbl_ingresos      = self._tarjeta(self.frame_tarjetas, "Ingresos",       COLOR_INGRESO)
        self.lbl_egresos       = self._tarjeta(self.frame_tarjetas, "Egresos",        COLOR_EGRESO)
        self.lbl_saldo_final   = self._tarjeta(self.frame_tarjetas, "Saldo final",    COLOR_NEUTRO)

        # ── Desglose por categoría ──
        tk.Label(self, text="Desglose por categoría", font=("Arial", 12, "bold"),
                 bg="#f0f2f5", fg=COLOR_NEUTRO).pack(anchor="w", padx=24, pady=(10, 4))

        frame_tabla = tk.Frame(self, bg="#f0f2f5")
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        cols = ("Categoría", "Tipo", "Total")
        self.tree = ttk.Treeview(frame_tabla, columns=cols, show="headings", height=14)

        for col in cols:
            self.tree.heading(col, text=col)

        self.tree.column("Categoría", width=260, anchor="w")
        self.tree.column("Tipo",      width=100, anchor="center")
        self.tree.column("Total",     width=140, anchor="e")

        self.tree.tag_configure("ingreso", foreground=COLOR_INGRESO)
        self.tree.tag_configure("egreso",  foreground=COLOR_EGRESO)

        scroll = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

    def _tarjeta(self, parent, titulo: str, color: str) -> tk.Label:
        """Crea una tarjeta de resumen y retorna el label del monto."""
        card = tk.Frame(parent, bg="white", bd=0, relief="flat",
                        highlightbackground="#dde1e7", highlightthickness=1)
        card.pack(side="left", expand=True, fill="both", padx=(0, 12))

        tk.Label(card, text=titulo, font=("Arial", 9),
                 bg="white", fg="#7f8c8d").pack(anchor="w", padx=16, pady=(14, 0))

        lbl_monto = tk.Label(card, text="$ -", font=("Arial", 18, "bold"),
                             bg="white", fg=color)
        lbl_monto.pack(anchor="w", padx=16, pady=(2, 14))

        return lbl_monto

    # ─── Carga de datos ───────────────────────────────────────

    def _cargar_meses(self):
        self._meses = obtener_meses()

        if not self._meses:
            self.combo_mes.set("Sin meses registrados")
            return

        etiquetas = [f"{MESES_NOMBRES[m['mes']]} {m['año']}" for m in self._meses]
        self.combo_mes["values"] = etiquetas

        mes_actual = obtener_mes_actual()
        if mes_actual:
            ids = [m["id"] for m in self._meses]
            if mes_actual["id"] in ids:
                self.combo_mes.current(ids.index(mes_actual["id"]))
            else:
                self.combo_mes.current(0)
        else:
            self.combo_mes.current(0)

        self._actualizar_datos()

    def _on_mes_seleccionado(self, _event=None):
        self._actualizar_datos()

    def _actualizar_datos(self):
        idx = self.combo_mes.current()
        if idx < 0 or not self._meses:
            return

        self.mes_actual = self._meses[idx]
        resumen  = obtener_resumen_mes(self.mes_actual["id"])
        desglose = obtener_desglose_por_categoria(self.mes_actual["id"])

        si            = self.mes_actual["saldo_inicial"]
        total_ing     = resumen["total_ingresos"]
        total_egr     = resumen["total_egresos"]
        saldo_final   = si + total_ing - total_egr

        # Actualizar tarjetas
        self.lbl_saldo_inicial.config(text=f"$ {si:,.2f}")
        self.lbl_ingresos.config(text=f"$ {total_ing:,.2f}")
        self.lbl_egresos.config(text=f"$ {total_egr:,.2f}")

        color_saldo = COLOR_INGRESO if saldo_final >= 0 else COLOR_EGRESO
        self.lbl_saldo_final.config(text=f"$ {saldo_final:,.2f}", fg=color_saldo)

        # Actualizar tabla de desglose
        self.tree.delete(*self.tree.get_children())
        for fila in desglose:
            self.tree.insert("", "end", values=(
                fila["categoria"],
                fila["tipo"].capitalize(),
                f"$ {fila['total']:,.2f}",
            ), tags=(fila["tipo"],))

    def _exportar_excel(self):
        if not self.mes_actual:
            messagebox.showwarning("Sin mes", "No hay ningún mes seleccionado.")
            return

        nombre_mes = MESES_NOMBRES[self.mes_actual["mes"]]
        año = self.mes_actual["año"]

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"resumen_{nombre_mes}_{año}.xlsx",
            title="Guardar exportación"
        )
        if not ruta:
            return

        resumen = obtener_resumen_mes(self.mes_actual["id"])
        movimientos = obtener_movimientos(self.mes_actual["id"])
        si = self.mes_actual["saldo_inicial"]
        saldo_final = si + resumen["total_ingresos"] - resumen["total_egresos"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{nombre_mes} {año}"

        # ── Estilos ──
        estilo_titulo = Font(bold=True, size=14)
        estilo_header = Font(bold=True, color="FFFFFF")
        fill_header   = PatternFill("solid", fgColor="2C3E50")
        fill_ingreso  = PatternFill("solid", fgColor="D5F5E3")
        fill_egreso   = PatternFill("solid", fgColor="FADBD8")
        alinear_der   = Alignment(horizontal="right")

        # ── Título ──
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Resumen — {nombre_mes} {año}"
        ws["A1"].font = estilo_titulo
        ws["A1"].alignment = Alignment(horizontal="center")

        # ── Bloque resumen ──
        datos_resumen = [
            ("Saldo inicial",  si),
            ("Ingresos",       resumen["total_ingresos"]),
            ("Egresos",        resumen["total_egresos"]),
            ("Saldo final",    saldo_final),
        ]
        for i, (etiqueta, valor) in enumerate(datos_resumen, start=3):
            ws[f"A{i}"] = etiqueta
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = valor
            ws[f"B{i}"].number_format = '#,##0.00'
            ws[f"B{i}"].alignment = alinear_der

        # ── Encabezados de la tabla ──
        fila_header = 8
        encabezados = ["Fecha", "Categoría", "Subcategoría", "Descripción", "Tipo", "Monto"]
        for col, texto in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_header, column=col, value=texto)
            celda.font = estilo_header
            celda.fill = fill_header
            celda.alignment = Alignment(horizontal="center")

        # ── Movimientos ──
        for fila, mov in enumerate(movimientos, start=fila_header + 1):
            ws.cell(row=fila, column=1, value=mov["fecha"])
            ws.cell(row=fila, column=2, value=mov["categoria"])
            ws.cell(row=fila, column=3, value=mov["subcategoria"] or "")
            ws.cell(row=fila, column=4, value=mov["descripcion"] or "")
            ws.cell(row=fila, column=5, value=mov["tipo"].capitalize())
            celda_monto = ws.cell(row=fila, column=6, value=mov["monto"])
            celda_monto.number_format = '#,##0.00'
            celda_monto.alignment = alinear_der

            fill = fill_ingreso if mov["tipo"] == "ingreso" else fill_egreso
            for col in range(1, 7):
                ws.cell(row=fila, column=col).fill = fill

        # ── Ancho de columnas ──
        anchos = [14, 20, 20, 30, 12, 14]
        for col, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

        wb.save(ruta)
        messagebox.showinfo("Exportación exitosa", f"Archivo guardado en:\n{ruta}")  